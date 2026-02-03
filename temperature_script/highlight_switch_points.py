from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from highlight_registry import HighlightRegistry, HighlightPoint

class HighlightSwitchPoints:
    def __init__(self, file_path: str, config: dict, registry: HighlightRegistry):
        self.file_path = file_path
        self.config = config
        self.registry = registry 
        
        # Load workbook (data_only=True to get values, formulas evaluated)
        self.wb = load_workbook(file_path, data_only=True)
        self.ws = self.wb.active

        # Create fills from config
        self.green_fill = PatternFill(
            start_color=config.highlightColors.green,
            end_color=config.highlightColors.green,
            fill_type="solid"
        )

        self.yellow_fill = PatternFill(
            start_color=config.highlightColors.yellow,
            end_color=config.highlightColors.yellow,
            fill_type="solid"
        )
        
        self.protected_headers = config.protectedHeaders

    def highlight_switch_points(self):
        """
        Highlights switch points using read_only + write_only mode for maximum performance.
        Open (1→0) as green, Closed (0→1) as yellow.
        """
        # 1. Open workbook in read_only mode
        data = list(self.ws.iter_rows(values_only=True))
        if len(data) < 2:
            return

        headers = data[0]
        pressure_col_idx = self.config.pressureCol - 1  # zero-based

        # Identify digital columns
        digital_indices = [
            i for i, header in enumerate(headers) if header not in self.protected_headers
        ]

        # -------------------------------
        # 2. Collect switch points in memory
        # -------------------------------
        highlights_to_apply = []
        prev_vals = {col_idx: data[1][col_idx] for col_idx in digital_indices}

        for row_idx in range(2, len(data)):
            row_values = data[row_idx]
            
            for col_idx in digital_indices:
                curr_val = row_values[col_idx]
                pressure = row_values[pressure_col_idx]

                if prev_vals[col_idx] == 1 and curr_val == 0:
                    highlights_to_apply.append(
                        (row_idx + 1, col_idx + 1, self.green_fill, True, curr_val, pressure, headers[col_idx])
                    )
                
                elif prev_vals[col_idx] == 0 and curr_val == 1:
                    highlights_to_apply.append(
                        (row_idx + 1, col_idx + 1, self.yellow_fill, False, curr_val, pressure, headers[col_idx])
                    )

                prev_vals[col_idx] = curr_val

        # -------------------------------
        # 3. Apply highlights and update registry
        # -------------------------------
        for r, c, fill, is_open, val, pressure, header in highlights_to_apply:
            
            self.ws.cell(row=r, column=c).fill = fill
            point = HighlightPoint(r, c, is_open, header, val, pressure)
            self.registry.record_event(point)

        self.wb.save(self.file_path)