from openpyxl import load_workbook
from typing import Optional

class CreateCleanedSheet:
    def __init__(self, file_path: str, highlighter):
        self.file_path = file_path
        self.highlighter = highlighter
        self.wb = load_workbook(file_path)
        self.ws = self.wb.active

    @staticmethod
    def is_highlighted(cell) -> bool:
        return cell.fill and cell.fill.fill_type and cell.fill.start_color.rgb in ("FF00FF00", "FFFFFF00")

    def create_cleaned_sheet(self, sheet_name: str = "Cleaned"):
        if sheet_name in self.wb.sheetnames:
            del self.wb[sheet_name]

        ws_out = self.wb.create_sheet(sheet_name)

        # Copy header
        for col in range(1, self.ws.max_column + 1):
            ws_out.cell(row=1, column=col).value = self.ws.cell(row=1, column=col).value

        out_row = 2

        for row in range(2, self.ws.max_row + 1):
            row_cells = [self.ws.cell(row=row, column=col) for col in range(1, self.ws.max_column + 1)]
            if any(self.is_highlighted(c) for c in row_cells):
                for col_idx, cell in enumerate(row_cells, start=1):
                    new_cell = ws_out.cell(row=out_row, column=col_idx)
                    new_cell.value = cell.value
                    new_cell.fill = cell.fill
                out_row += 1

        self.wb.save(self.file_path)

