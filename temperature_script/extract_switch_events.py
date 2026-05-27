import pandas as pd
from openpyxl import load_workbook
from highlight_registry import HighlightRegistry


class ExtractSwitchEvents:
    def __init__(self, file_path: str, config: dict, registry: HighlightRegistry):
        self.file_path = file_path
        self.config = config
        self.registry = registry

        self.wb = load_workbook(file_path)
        self.ws = self.wb.active

        self.digital_start_col: int = config.digitalStartCol
        self.pressure_col: int = config.pressureCol
        self.protected_headers = config.protectedHeaders

    def create_switch_events_sheet(self, sheet_name: str = "SwitchEvents"):

        headers = [
            self.ws.cell(row=1, column=col).value
            for col in range(1, self.ws.max_column + 1)
        ]

        sessions_by_col = self.registry.get_sessions_by_column()
        
        digital_cols = list(sessions_by_col.keys()) 

        data = {header: [] for header in headers}

        pressure_per_row = {
            point.row: point.pressure
            for point in self.registry.lookup.values()
        }

        complete_event_count = min(len(sessions) for sessions in sessions_by_col.values())
     
        for event_idx in range(complete_event_count):

            event_rows = set()
            
            for sessions in sessions_by_col.values():
                
                session = sessions[event_idx]
                event_rows.add(session.open_point.row)
                event_rows.add(session.close_point.row)

            for row in sorted(event_rows):
                
                row_values = []

                for col_idx, header in enumerate(headers, start=1):

                    if header in self.protected_headers:

                        value = self.ws.cell(row=row, column=col_idx).value

                    elif col_idx in digital_cols:

                        point = self.registry.lookup.get((row, col_idx))
                        value = point.value if point else None

                    elif col_idx == self.pressure_col:
                        value = pressure_per_row.get(row)

                    else:
                        value = None

                    row_values.append(value)

                for idx, val in enumerate(row_values):
                    data[headers[idx]].append(val)


            diff_row = [None] * len(headers)
            diff_row[0] = "Differential"
            percent_diff_row = [None] * len(headers)
            percent_diff_row[0] = "Percent differential (%)"
            
            for col_idx in digital_cols:
                
                session = sessions_by_col[col_idx][event_idx]
                open_pressure = pressure_per_row.get(session.open_point.row)
                close_pressure = pressure_per_row.get(session.close_point.row)

                if open_pressure is not None and close_pressure is not None:
                    diff = open_pressure - close_pressure
                    diff_row[col_idx - 1] = diff
                    percent_diff_row[col_idx - 1] = (diff / open_pressure) * 100

            for idx, val in enumerate(diff_row):              
                data[headers[idx]].append(val)
            
            for idx, val in enumerate(percent_diff_row):                
                data[headers[idx]].append(val)

            for header in headers:                
                data[header].append(None)

        df = pd.DataFrame(data)

        with pd.ExcelWriter(self.file_path, engine="openpyxl", mode="a") as writer:
            if sheet_name in writer.book.sheetnames:
                idx = writer.book.sheetnames.index(sheet_name)
                writer.book.remove(writer.book.worksheets[idx])

            df.to_excel(writer, sheet_name=sheet_name, index=False)
