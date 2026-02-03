import pandas as pd
from openpyxl import load_workbook
from highlight_registry import HighlightRegistry


class ExtractSwitchEvents:
    def __init__(self, file_path: str, config: dict, registry: HighlightRegistry):
        self.file_path = file_path
        self.config = config
        self.registry = registry

        self.wb = load_workbook(file_path)
        self.ws = self.wb.active

        self.digital_start_col: int = config.digitalStartCol
        self.pressure_col: int = config.pressureCol
        self.protected_headers = config.protectedHeaders

    def create_switch_events_sheet(self, sheet_name: str = "SwitchEvents"):
        """
        Builds a SwitchEvents sheet by iterating switch SESSIONS directly
        instead of reconstructing events from rows.
        """

        # -----------------------------
        # 1. Headers & digital columns
        # -----------------------------
        headers = [
            self.ws.cell(row=1, column=col).value
            for col in range(1, self.ws.max_column + 1)
        ]

        digital_cols = list(range(self.digital_start_col, self.ws.max_column + 1))

        # Output structure for pandas
        data = {header: [] for header in headers}

        sessions_by_col = self.registry.get_sessions_by_column()

        # -----------------------------
        # 2. Determine number of complete events
        #    (event index = same index across columns)
        # -----------------------------
        complete_event_count = min(
            len([s for s in sessions if s.is_complete])
            for sessions in sessions_by_col.values()
        )

        # -----------------------------
        # 3. Process each event (SESSION-DRIVEN)
        # -----------------------------
        for event_idx in range(complete_event_count):

            # Collect all rows involved in this event
            event_rows = set()

            for sessions in sessions_by_col.values():
                session = sessions[event_idx]
                event_rows.add(session.open_point.row)
                event_rows.add(session.close_point.row)

            # ---------------------------------
            # 3a. Write event rows
            # ---------------------------------
            for row in sorted(event_rows):
                row_values = []

                for col_idx, header in enumerate(headers, start=1):

                    if header in self.protected_headers:
                        value = self.ws.cell(row=row, column=col_idx).value

                    elif col_idx in digital_cols:
                        # Only show 0/1 at actual switch points
                        value = None
                        for session in sessions_by_col.get(col_idx, []):
                            if session.open_point.row == row:
                                value = session.open_point.value
                                break
                            if session.close_point and session.close_point.row == row:
                                value = session.close_point.value
                                break

                    else:
                        value = None

                    row_values.append(value)

                for idx, val in enumerate(row_values):
                    data[headers[idx]].append(val)

            # ---------------------------------
            # 3b. Differential row
            # ---------------------------------
            diff_row = [None] * len(headers)
            diff_row[0] = "Differential"

            for col_idx in digital_cols:
                session = sessions_by_col[col_idx][event_idx]

                open_pressure = self.ws.cell(
                    row=session.open_point.row,
                    column=self.pressure_col
                ).value

                close_pressure = self.ws.cell(
                    row=session.close_point.row,
                    column=self.pressure_col
                ).value

                if open_pressure is not None and close_pressure is not None:
                    diff_row[col_idx - 1] = open_pressure - close_pressure

            for idx, val in enumerate(diff_row):
                data[headers[idx]].append(val)

            # ---------------------------------
            # 3c. Blank separator row
            # ---------------------------------
            for header in headers:
                data[header].append(None)

        # -----------------------------
        # 4. Write to Excel
        # -----------------------------
        df = pd.DataFrame(data)

        with pd.ExcelWriter(self.file_path, engine="openpyxl", mode="a") as writer:
            if sheet_name in writer.book.sheetnames:
                idx = writer.book.sheetnames.index(sheet_name)
                writer.book.remove(writer.book.worksheets[idx])

            df.to_excel(writer, sheet_name=sheet_name, index=False)
"""
        # Stores all the rows where the switch opens or closes, and sorts them
        all_rows_set = set()
        
        for col_sessions in self.registry.get_sessions_by_column().values():
            
            for session in col_sessions:
                
                all_rows_set.add(session.open_point.row)
                
                if session.close_point:
                    all_rows_set.add(session.close_point.row)

        all_rows = sorted(all_rows_set)

        # Stores the headers in an array, maybe could use a set as they're unique
        headers = [self.ws.cell(row=1, column=col).value for col in range(1, self.ws.max_column + 1)]

        # Initialize a dictionary of lists for each column
        # TMA-7.0 NBR: [],
        data = {header: [] for header in headers}

        # this is a dictionary to quickly get a switch value (1/0) at a specified (row,col), which is a key
        # (97, 5): 0 where 97 is a row and 5 is a column (switch, TMA 7.0)
        # Avoids the .cell() call to look for the cell everytime
        row_col_values = {}
        
        # list of columns that hold the digital values of the switches
        digital_cols = list(range(self.digital_start_col, self.ws.max_column + 1))

        open_rows_per_col = {col: set() for col in digital_cols}
        close_rows_per_col = {col: set() for col in digital_cols}
        
        for col_idx, col_sessions in self.registry.get_sessions_by_column().items():

            for session in col_sessions:
                
                # store the value for later sheet output
                row_col_values[(session.open_point.row, col_idx)] = session.open_point.value
                if session.close_point:
                    row_col_values[(session.close_point.row, col_idx)] = session.close_point.value

                # track which rows are open/closed for differential logic
                open_rows_per_col[col_idx].add(session.open_point.row)
                
                if session.close_point: 
                    close_rows_per_col[col_idx].add(session.close_point.row)

        # --- 5. Fill data dictionary row by row ---
        event_rows = []
        event_columns = set()
        closed_columns_in_event = set()

        for row in all_rows:
            row_data = []
            
            for col_idx, header in enumerate(headers, start=1):
                if header in self.protected_headers:
                    value = self.ws.cell(row=row, column=col_idx).value
                elif col_idx in digital_cols:
                    # Only show 0/1 for switch points, blank otherwise
                    value = row_col_values.get((row, col_idx), None)
                else:
                    value = None
                row_data.append(value)

            for idx, val in enumerate(row_data):
                data[headers[idx]].append(val)

            # --- 5. Event tracking for differential ---
            event_rows.append(row)
           
            for col_idx in digital_cols:
                
                if row in open_rows_per_col[col_idx]:
                    event_columns.add(col_idx)
                elif row in close_rows_per_col[col_idx] and col_idx in event_columns:
                    closed_columns_in_event.add(col_idx)

            if event_columns and event_columns == closed_columns_in_event:
                
                diff_row_data = [None] * len(headers)
                diff_row_data[0] = "Differential"
                
                for col_idx in digital_cols:
                    open_vals = [
                        self.ws.cell(r, self.pressure_col).value
                        for r in event_rows if r in open_rows_per_col[col_idx]
                    ]
                    closed_vals = [
                        self.ws.cell(r, self.pressure_col).value
                        for r in event_rows if r in close_rows_per_col[col_idx]
                    ]
                    if open_vals and closed_vals:
                        diff_row_data[col_idx - 1] = max(open_vals) - min(closed_vals)

                for idx, val in enumerate(diff_row_data):
                    data[headers[idx]].append(val)

                # Blank row after differential
                for header in headers:
                    data[header].append(None)

                # Reset for next event block
                event_rows.clear()
                event_columns.clear()
                closed_columns_in_event.clear()
"""