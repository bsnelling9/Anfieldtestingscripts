from openpyxl import load_workbook
from highlight_registry import HighlightRegistry

def export_registry_in_excel(file_path: str, registry: HighlightRegistry, sheet_name="RegistryExport"):
 
    wb = load_workbook(file_path)

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    ws.append(["Row", "Column", "Header", "Is Open Circuit", "Value"])
    for col_idx, sessions in registry.columns.items():
        for session in sessions:
            ws.append([
                session.open_point.row,
                session.open_point.col,
                session.open_point.header,
                True,
                session.open_point.value
            ])
            if session.close_point:
                ws.append([
                    session.close_point.row,
                    session.close_point.col,
                    session.close_point.header,
                    False,
                    session.close_point.value
            ])
    wb.save(file_path)
    print(f"Registry exported in '{sheet_name}' sheet of {file_path}")
